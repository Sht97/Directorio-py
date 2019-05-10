import tkinter as tk
from tkinter import filedialog, messagebox, Menu, StringVar, Entry, Label, Button, Text, Scrollbar
from tkinter.ttk import Treeview
import re
import openpyxl


def open_file():
    global save_path
    old_save_path = save_path
    save_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if save_path == "" or save_path == "()":
        messagebox.showinfo(message="Debe seleccionar un archivo", title="Error")
        save_path = old_save_path
    a = open('Config/path.txt', 'w')
    a.write(save_path)
    a.close()
    delete_all()
    global data
    try:
        data = list(openpyxl.load_workbook(save_path)["Hoja 1"])
    except:
        data = []
    append()


def new():
    global save_path
    old_save_path = save_path
    save_path = filedialog.asksaveasfilename(filetypes=[("Excel files", "*.xlsx *.xls")], defaultextension=".xlsx")
    print(save_path)
    if save_path == "" or save_path == "()":
        messagebox.showinfo(message="No se ha podido crear el nuevo directorio", title="Error")
        save_path = old_save_path
    else:
        a = open('Config/path.txt', 'w')
        a.write(save_path)
        a.close()
        delete_all()
        aux.set("new")
        update()


def update():
    global save_path
    wb = openpyxl.Workbook()
    hoja = wb.active
    hoja.title = "Hoja 1"
    fila = 1  # Fila donde empezamos
    for i in tv.get_children():
        hoja.cell(column=1, row=fila, value=tv.item(i)['text'])
        hoja.cell(column=2, row=fila, value=tv.item(i)['values'][0])
        hoja.cell(column=3, row=fila, value=tv.item(i)['values'][1])
        hoja.cell(column=4, row=fila, value=tv.item(i)['values'][2])
        hoja.cell(column=5, row=fila, value=tv.item(i)['values'][3])
        fila += 1
    wb.save(filename=save_path)
    if aux.get() != "new":
        messagebox.showinfo("Guardar", "Se ha actualizado la información.")
    aux.set(" ")


def find():
    if nameToFind.get() != "":
        btnBack.place(x=250, y=300)
        btnAdd.place_forget()
        btnEdit.place_forget()
        btnDelete.place_forget()
        selections = []
        for i in tv.get_children():
            if nameToFind.get() in tv.item(i)['values'] or tv.item(i)['text'] == nameToFind.get():
                selections.append(i)
        tv.selection_set(selections)
        if selections == []:
            messagebox.showinfo("Error", "El contacto no se encuentra en la lista.")
            back()
    else:
        messagebox.showinfo("Error", "No ha ingresado ningún texto.")


def back():
    btnBack.place_forget()
    btnAdd.place(x=100, y=300)
    btnEdit.place(x=250, y=300)
    btnDelete.place(x=400, y=300)
    nameToFind.set("")
    tv.selection_set()


def add():
    root2.deiconify()
    btn.place_forget()
    btn2.place(x=60, y=230)


def edit():
    if tv.focus():
        selected = tv.item(tv.focus())
        name.set(selected['text'])
        lastName.set((selected['values'][0]))
        tel.set((selected['values'][1]))
        cel.set((selected['values'][2]))
        email.set((selected['values'][3]))
        btn2.place_forget()
        btn.place(x=60, y=230)
        root2.deiconify()
    else:
        messagebox.showinfo("Error", "Ninguna persona ha sido seleccionada.")


def add_edit():
    if (name.get() == "" or lastName.get() == "" or cel.get() == "" or tel.get() == "" or email.get() == ""):
        messagebox.showinfo("Error", "Debe llenar todos los datos.")
    else:
        if validate():
            tv.item(tv.focus(), text=name.get(), values=[lastName.get(), tel.get(), cel.get(), email.get()])
            root2.withdraw()
            btn.place_forget()
            cleanBox()


def delete_all():
    for i in tv.get_children():
        tv.delete(i)


def delete():
    if tv.focus():
        tv.delete(tv.focus())
    else:
        messagebox.showinfo("Error", "Ninguna persona ha sido seleccionada.")


def load_table():
    if (name.get() == "" or lastName.get() == "" or cel.get() == "" or tel.get() == "" or email.get() == ""):
        messagebox.showinfo("Error", "Debe llenar todos los datos.")
    else:
        if validate():
            tv.insert('', 'end', text=name.get(), values=(lastName.get(), tel.get(), cel.get(), email.get()))
            root2.withdraw()
            cleanBox()


def append():
    for i in range(len(data)):
        try:
            tv.insert('', 'end', text=data[i][0].value,
                      values=(data[i][1].value, data[i][2].value, data[i][3].value, data[i][4].value))
        except:
            messagebox.showinfo("Aviso", "Directorio vacío")


def cancel():
    root2.withdraw()
    cleanBox()


def validate():
    validateVar = 0
    patronName = "(([A-Za-zñÑáéíóúÁÉÍÓÚ]+(\\s))?[A-Za-zñÑáéíóúÁÉÍÓÚ])+$"
    patronTel = "([0-9]{1,4})+$"
    patronEmail = "([A-Za-zñÑ]+|(\\d))+((.+|_+)([a-z]+|(\\d+)))?@[A-Za-z]+.com$"
    if re.match(patronName, name.get()):
        validateVar += 1
    else:
        messagebox.showinfo("Error", 'El campo "Nombre" no cumple con el formato.')
    if re.match(patronName, lastName.get()):
        validateVar += 1
    else:
        messagebox.showinfo("Error", 'El campo "Apellido" no cumple con el formato.')
    if re.match(patronTel, tel.get()):
        validateVar += 1
    else:
        messagebox.showinfo("Error", 'El campo "Teléfono" no cumple con el formato.')
    if re.match(patronTel, cel.get()):
        validateVar += 1
    else:
        messagebox.showinfo("Error", 'El campo "Celular" no cumple con el formato.')
    if re.match(patronEmail, email.get()):
        validateVar += 1
    else:
        messagebox.showinfo("Error", 'El campo "Correo" no cumple con el formato.')

    if (validateVar == 5):
        return True
    else:
        return False


def cleanBox():
    name.set("")
    lastName.set("")
    tel.set("")
    cel.set("")
    email.set("")


f = open('Config/path.txt', 'r')
save_path = f.read()
f.close()
try:
    data = list(openpyxl.load_workbook(save_path)["Hoja 1"])
except:
    data = []

root = tk.Tk()
root2 = tk.Toplevel(root)
menubar = Menu(root)
text = Text(root)
name = StringVar()
nameToFind = StringVar()
lastName = StringVar()
tel = StringVar()
email = StringVar()
cel = StringVar()
aux = StringVar()
bgColor = "#BAFADC"
bgColor2 = "#DDFFBE"

root.title("Agenda")
root.geometry("625x350")
root.configure(background=bgColor)

root2.overrideredirect(1)
root2.title("Contacto")
root2.geometry("280x280")
root2.configure(background=bgColor2)
root2.withdraw()

filemenu = Menu(menubar, tearoff=0)
filemenu.configure(background="white")
filemenu.add_command(label="Nuevo", command=new)
filemenu.add_command(label="Abrir", command=open_file)
filemenu.add_command(label="Guardar", command=update)
filemenu.add_separator()
filemenu.add_command(label="Salir", command=root.quit)
menubar.add_cascade(label="Archivo", menu=filemenu)
filemenu.add_separator()

Label(root2, text="AGREGAR CONTACTO", bg=bgColor2, fg="black").place(x=80, y=20)
Entry(root2, textvariable=name).place(x=120, y=60)
Label(root2, text="Nombre: ", bg=bgColor2, fg="black").place(x=30, y=60)
Entry(root2, textvariable=lastName).place(x=120, y=90)
Label(root2, text="Apellido: ", bg=bgColor2, fg="black").place(x=30, y=90)
Entry(root2, textvariable=tel).place(x=120, y=120)
Label(root2, text="Teléfono: ", bg=bgColor2, fg="black").place(x=30, y=120)
Entry(root2, textvariable=cel).place(x=120, y=150)
Label(root2, text="Celular: ", bg=bgColor2, fg="black").place(x=30, y=150)
Entry(root2, textvariable=email).place(x=120, y=180)
Label(root2, text="Email: ", bg=bgColor2, fg="black").place(x=30, y=180)
btn = Button(root2, text="Guardar", command=add_edit, bg="white", fg="black", width=8)
btn2 = Button(root2, text="Guardar", command=load_table, bg="white", fg="black", width=8)
Button(root2, text="Cancelar", command=cancel, bg="white", fg="black", width=8).place(x=150, y=230)

tv = Treeview()
tv['columns'] = ('Apellido', 'Telefono Fijo', 'Celular', 'Correo electronico')
tv.heading("#0", text='  Nombre', anchor='w')
tv.column("#0", anchor="w", width=100)
tv.heading('Apellido', text='  Apellido', anchor='w')
tv.column('Apellido', anchor="w", width=100)
tv.heading('Telefono Fijo', text='Teléfono Fijo')
tv.column('Telefono Fijo', anchor='w', width=100)
tv.heading('Celular', text='Celular')
tv.column('Celular', anchor='w', width=120)
tv.heading('Correo electronico', text='Correo electrónico')
tv.column('Correo electronico', anchor='w', width=180)
tv.grid(sticky=("N", "S", "W", "E"))
tv.place(x=10, y=50)
scrollBar = Scrollbar(root, orient="vertical", command=tv.yview)
scrollBar.place(x=600, y=51, height=225)
tv.configure(yscrollcommand=scrollBar.set)

append()

Entry(root, textvariable=nameToFind, width=40).place(x=130, y=10)
Button(root, text="Buscar", command=find, bg="white", fg="black", width=10).place(x=420, y=10)
btnBack = Button(root, text="Regresar", command=back, bg="white", fg="black", width=10)

btnAdd = Button(root, text="Agregar", command=add, bg="white", fg="black", width=10)
btnAdd.place(x=100, y=300)
btnEdit = Button(root, text="Editar", command=edit, bg="white", fg="black", width=10)
btnEdit.place(x=250, y=300)
btnDelete = Button(root, text="Eliminar", command=delete, bg="white", fg="black", width=10)
btnDelete.place(x=400, y=300)

root.config(menu=menubar)
root.mainloop()
